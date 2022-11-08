import streamlit as st
import openpyxl  as op
import math

st.image('SPSA.png')

def Y1():
    st.title('ущерб нанесённый пожаром объектам строительства, руб.')
    Number = int(st.number_input("Введите число объектов строительства на которых пожаром была поврежденна или уничтоженна их площадь ", step = 1))
    Y34 = []
    def Y_34_price(klfg):
        asas = klfg + 1
        st.subheader(f'Рассчёт ущерба на {asas} объекте')
        st.write('Расчет ущерба в результате уничтожения пожаром площади объекта строительства')
        S3 =st.number_input('Введите повреждённую пожаром площадь i-го объекта строительства кв. м.',min_value = 0.01, key=str(klfg)+'5345343846335')
        S1 =st.number_input('Введите уничтоженную пожаром площадь i-го объекта строительства кв. м.',min_value = 0.01, key=str(klfg)+'584674745')
        #приложеие А
        tabl_pril_A = {
            'ОБЪЕКТЫ ЗДРАВООХРАНЕНИЯ':1,
            'ОБЪЕКТЫ СПОРТИВНОГО НАЗНАЧЕНИЯ': 2,
            'ОБЪЕКТЫ ОБРАЗОВАНИЯ': 3,
            'ОБЪЕКТЫ КУЛЬТУРЫ': 4,
            'ОБЪЕКТЫ ТОРГОВЛИ': 5,
            'ОБЪЕКТЫ ОБЩЕСТВЕННОГО ПИТАНИЯ' : 6,
            'ОБЪЕКТЫ БЫТОВОГО И КОММУНАЛЬНОГО НАЗНАЧЕНИЯ': 7, 
            'ЖИЛЫЕ ДОМА И ОБЪЕКТЫ ПРИУСАДЕБНОГО ХОЗЯЙСТВА': 8,
            'ОБЪЕКТЫ ДЛЯ ВРЕМЕННОГО ПРЕБЫВАНИЯ ЛЮДЕЙ': 9,
            'ПРОИЗВОДСТВЕННЫЕ ЗДАНИЯ': 10,
            'ОБЪЕКТЫ ЛЕСНОЙ, ДЕРЕВООБРАБАТЫВАЮЩЕЙ ЦЕЛЛЮЛОЗНО-БУМАЖНОЙ ПРОМЫШЛЕННОСТИ': 11,
            'ОБЪЕКТЫ ЛЕГКОЙ ПРОМЫШЛЕННОСТИ': 12,
            'ОБЪЕКТЫ ГАЗОВОЙ ПРОМЫШЛЕННОСТИ': 13,
            'ОБЪЕКТЫ ПРОМЫШЛЕННОСТИ СТРОИТЕЛЬНЫХ МАТЕРИАЛОВ, ДЕТАЛЕЙ И КОНСТРУКЦИЙ': 14,
            'ОБЪЕКТЫ ЧЕРНОЙ МЕТАЛЛУРГИИ': 15,
            'ОБЪЕКТЫ ХИМИЧЕСКОЙ ПРОМЫШЛЕННОСТИ': 16,
            'ОБЪЕКТЫ НЕФТЯНОЙ ПРОМЫШЛЕННОСТИ': 17,
            'ОБЪЕКТЫ МЕДИЦИНСКОЙ И МИКРОБИОЛОГИЧЕСКОЙ ПРОМЫШЛЕННОСТИ' : 18,
            'ОБЪЕКТЫ МАШИНОСТРОИТЕЛЬНОЙ ПРОМЫШЛЕННОСТИ': 19,
            'ОБЪЕКТЫ УГОЛЬНОЙ, СЛАНЦЕВОЙ, ТОРФЯНОЙ, ГОРНОРУДНОЙ И ГОРНОДОБЫВАЮЩЕЙ ПРОМЫШЛЕННОСТИ': 20,
            'Объекты полиграфической и кинематографической промышленности': 21,
            'ОБЪЕКТЫ ПИЩЕВОЙ, МУКОМОЛЬНОЙ И КОМБИКОРМОВОЙ ПРОМЫШЛЕННОСТИ': 22,
            'ОБЪЕКТЫ ЖИВОТНОВОДСТВА': 23,
            'ОБЪЕКТЫ РАСТЕНИЕВОДСТВА': 24,
            'объекты строительства складского назначения':25,
            'Отсутствуют сведения об объектах того же функционального назначения': 0,        
        }   
        txt_A = st.selectbox(
            'Выберите объект, для определения нормативной капитальности',
            tabl_pril_A.keys(), key=str(klfg)+'5934534595' 
        )

        vibor_A = tabl_pril_A.get(txt_A)
    #начали цену 
        if vibor_A == 0:
            C_social_price_no_infl = st.number_input('Введите общую стоимость объекта строительства, руб.', key=str(klfg)+'5986538605')
            S_OB_build = st.number_input('Введите общую площадь объекта строительства, руб.', min_value = 0.01, key=str(klfg)+'5925355')
            st.write('если разница между датой, на которую установлена стоимость объекта строительства, и датой пожара более одного года, то нажмите"дополнительный расчёт" ')
            if st.checkbox('дополнительный расчёт', key=str(klfg)+'5923738605'):
                N1 = st.number_input('год в который установленна кадастровая стоимость объекта', key=str(klfg)+'5925355',min_value = 2000,max_value = 2021, step=1,)
                N2 = st.number_input('Год в котором произошёл пожар', key=str(klfg)+'52644655',min_value = 2000,max_value = 2021, step=1,)
                pril_j = {
                    2000:1.202,
                    2001:1.1858,
                    2002:1.1506,
                    2003:1.1199,
                    2004:1.1174,
                    2005:1.1091,
                    2006:1.0900,
                    2007:1.1187,
                    2008:1.1328,
                    2009:1.088,
                    2010:1.0878,
                    2011:1.061,
                    2012:1.0658,
                    2013:1.0645,
                    2014:1.1136,
                    2015:1.1291,
                    2016:1.0538,
                    2017:1.0252,
                    2018:1.0427,
                    2019:1.0305,
                    2020:1.0491,
                    2021:1.0839,
                }
                dict_pril_j  = []
                for kf_in_pril_J in range(N1,N2+1):
                    vibor_in_pril_j = pril_j.get(kf_in_pril_J)
                    dict_pril_j.append(vibor_in_pril_j)
                k_infl = math.prod(dict_pril_j)
                C1 = (C_social_price_no_infl*k_infl)/S_OB_build
            else:
                C1 = C_social_price_no_infl/S_OB_build
            pril_i = {
                'одноквартирный жилой дом': 0.67,
                'многоквартирный жилой дом':0.35 ,
                'Объекты общественного назначения': 0.37,
                'Административные здания': 0.24,
                'Объекты складского, производственного назначения': 0.66,
                'Другие объекты строительства': 0.64,
            }
            tabl_i = st.selectbox(
                'Выберите тип объекта из предложенного списка',
                pril_i.keys(), key=str(klfg)+'567373534835336655'
            )
            kf_an_isn  = pril_i.get(tabl_i)
            Y1 = (kf_an_isn*C1*S3) + (C1*S1)
        else:
            book_A = op.open(f'А{vibor_A}.xlsx', read_only=True)
            start_A = book_A.active
            vc_A = dict()
            for ia, rowa in enumerate(range(1,start_A.max_row +1)):
                sub_rf_A = start_A[rowa][0].value
                vc_A[sub_rf_A] = ia

            sub_russia_A = st.selectbox(
                'Тип объекта данного объекта',
                vc_A, key=str(klfg)+'5234295'
            )
            for da,qa in enumerate(range(1,start_A.max_row +1)):
                if sub_russia_A == start_A[qa][0].value:
                    line_A = qa
    #здесь переменная А, которая для к3
            A = start_A[line_A][2].value
            st.write('Если присутствуют сведения о виде объекта строительства,но имеющего другие значения эксплуатационного параметра,нажмите "Индивидуальный расчёт"')
            if st.checkbox('Индивидуальный расчёт'):
                st.write('Выберите значение эксплуатационного параметра, того же функционального назначения,но, наиболее приближенный к вашему, меньше данного значения.')
                st.write('Если ваше значение меньше, чем данные в списке,то выберите наименьший параметр из предложенных и запишите его значение.')
                book_a = op.open(f'А{vibor_A}.xlsx', read_only=True)
                start_a = book_a.active
                vc_c1_a = dict()
                for aa,ba in enumerate(range(1,start_a.max_row +1)):
                    price_a = start_a[ba][0].value
                    vc_c1_a[price_a] = aa
                                    
                price_in_pril_a = st.selectbox(
                    'Здания и учреждения',
                    vc_c1_a, key=str(klfg)+'5697343655'
                )
                for ra,na in enumerate(range(1,start_a.max_row +1)):
                    if price_in_pril_a == start_a[na][0].value:
                        strk_a = na
                C_a = start_a[strk_a][1].value
                a = st.number_input('значение эксплуатационного параметра (меньший)',min_value = 1,step=1, key=str(klfg)+'5090905')
                st.write('Выберите значение эксплуатационного параметра, того же функционального назначения,но, наиболее приближенный к вашему, больше данного значения.')
                st.write('Если ваше значение больше, чем данные в списке,то выберите наибольший параметр из предложенных и запишите его значение.')
                book_c = op.open(f'А{vibor_A}.xlsx', read_only=True)
                start_c = book_c.active
                vc_c1_c = dict()
                for ac,bc in enumerate(range(1,start_c.max_row +1)):
                    price_c = start_c[bc][0].value
                    vc_c1_c[price_c] = ac
                                    
                price_in_pril_c = st.selectbox(
                    'Здания и учреждения',
                    vc_c1_c, key=str(klfg)+'5694810981655'
                )
                for rc,nc in enumerate(range(1,start_c.max_row +1)):
                    if price_in_pril_c == start_c[nc][0].value:
                        strk_c = nc
                C_c = start_c[strk_c][1].value 
                c = st.number_input('значение эксплуатационного параметра (больший)',min_value = 1,step=1, key=str(klfg)+'5800795')
                b = st.number_input('Теперь введите фактическое значение эксплуатационного параметра',min_value = 1,step=1, key=str(klfg)+'547546455')
                if (a < b < c):
                    C1 = C_c - ((c-b)*(C_c - C_a)/(c - a))
                elif a > b:
                    C1 = C_a
                elif c < b:
                    C1 = C_c
            else:
                C1 = start_A[line_A][1].value
    #начинаем коэф 1    
            T_Fi = st.number_input('Фактический срок службы i-го объекта в годах', step = 1, key=str(klfg)+'57565675',)
            st.write('Примечание.Если на объекте строительства была проведена реконструкция (модернизация),то фактический срок службы исчисляется, начиная с даты завершения реконструкции (модернизации).')
    #приложение Д
            file_XL_D = op.open('Д.xlsx', read_only=True)
            status_file_D = file_XL_D.active
            dict_pril_D = dict()
            for zd, wd in enumerate(range(1,status_file_D.max_row +1)):
                vid_build_D = status_file_D[wd][0].value
                dict_pril_D[vid_build_D] = zd

            vids_build_D = st.selectbox(
                'Выберите вид постройки для определения минимальный нормативный срок эксплуатации объекта',
                dict_pril_D, key=str(klfg)+'545'
            )
            for xd,md in enumerate(range(1,status_file_D.max_row +1)):
                if vids_build_D == status_file_D[md][0].value:
                    straka_D = md
            D = status_file_D[straka_D][1].value
            
            T1000 = D
            st.write('Примечание.Если вашего объекта нет в списке, то выберите пункт "Иные постройки" и введите его самостоятельно')
            if T1000 == 0:
                T_j = st.number_input('минимальный нормативный срок эксплуатации объекта строительства',min_value = 10, step = 1, key=str(klfg)+'57798795675',)
            else:
                T_j = T1000
            Fi_isn = (T_Fi/T_j)*100
            if Fi_isn <= 100:
    #прилодение Е
                file_XL_E = op.open('Е.xlsx', read_only=True)
                status_file_E = file_XL_E.active
                Fi_isn_1 = round(Fi_isn)
                for xe,me in enumerate(range(1,status_file_E.max_row +1)):
                    if Fi_isn_1 == status_file_E[me][0].value:
                        straka_E = me
                K1 = status_file_E[straka_E][1].value

            elif Fi_isn > 200:
                k1 = 96.08 + (T_Fi - T_j)*0.01
            else:
                k1 = 91.08 + (T_Fi - T_j)*0.05
    #закончили к1 
    #начали к4 и k2
            file_XL_G = op.open('Г.xlsx', read_only=True)
            status_file_G = file_XL_G.active
            dict_pril_G = dict()
            for zg, wg in enumerate(range(1,status_file_G.max_row +1)):
                vid_avto_G = status_file_G[wg][0].value
                dict_pril_G[vid_avto_G] = zg

            vids_avto_G = st.selectbox(
                'наименование территории субъекта рф',
                dict_pril_G, key=str(klfg)+'546565675'
            )
            for xg,mg in enumerate(range(1,status_file_G.max_row +1)):
                if vids_avto_G == status_file_G[mg][0].value:
                    straka_G = mg
            k4 = status_file_G[straka_G][1].value
            vocalbury_G_B = {
                1:51,
                2:66,
                3:31,
                4:85,
                5:85,
                6:59,
                7:59,
                8:63,
                9:60,
                10:52,
                11:61,
                12:2,
                13:2,
                14:3,
                15:3,
                16:3,
                17:32,
                18:33,
                19:76,
                20:76,
                21:76,
                22:76,
                23:76,
                24:76,
                25:76,
                26:62,
                27:34,
                28:67,
                29:35,
                30:68,
                31:64,
                32:36,
                33:69,
                34:53,
                35:53,
                36:53,
                37:70,
                38:70,
                39:70,
                40:70,
                41:70,
                42:78,
                43:78,
                44:78,
                45:78,
                46:65,
                47:79,
                48:79,
                49:79,
                50:79,
                51:80,
                52:4,
                53:4,
                54:4,
                55:4,
                56:4,
                57:54,
                58:13,
                59:14,
                60:15,
                61:55,
                62:6,
                63:6,
                64:16,
                65:17,
                66:71,
                67:71,
                68:71,
                69:71,
                70:7,
                71:18,
                72:77,
                73:77,
                74:77,
                75:77,
                76:77,
                77:72,
                78:38,
                79:19,
                80:19,
                81:45,
                82:20,
                83:8,
                84:9,
                85:21,
                86:81,
                87:81,
                88:81,
                89:81,
                90:81,
                91:22,
                92:30,
                93:10,
                94:10,
                95:39,
                96:11,
                97:73,
                98:74,
                99:40,
                100:23,
                101:41,
                102:37,
                103:12,
                104:56,
                105:24,
                106:42,
                107:43,
                108:82,
                109:82,
                110:82,
                111:82,
                112:82,
                113:82,
                114:46,
                115:25,
                116:26,
                117:27,
                118:75,
                119:28,
                120:47,
                121:47,
                122:47,
                123:44,
                124:50,
                125:29,
                126:83,
                127:5,
                128:5,
                129:5,
                130:5,
                131:48,
                132:48,
                133:84,
                134:49,
                135:49,
                136:86,
                137:86,
                138:57,
                139:58,
            }

            vibor_G_B = vocalbury_G_B.get(straka_G)
    #начали к2 
            book_b = op.open('Б.xlsx', read_only=True)
            start_b = book_b.active

            vc_b2 = dict()

            for kb,hb in enumerate(range(1,start_b.max_column)):
                institutions_b = start_b[1][hb].value
                vc_b2[institutions_b] = kb

            institution_b = st.selectbox(
                'Выберите объект строительства',
                vc_b2, key=str(klfg)+'85855'
            )

            for kb1,hb2 in enumerate(range(1,start_b.max_column)):
                if institution_b == start_b[1][hb2].value:
                    klfb = hb2 

            k2 = start_b[vibor_G_B][klfb].value  

    #закончили к2 и k4
    #начали к3
            vibor_fact_group = {
                'фактическая группа капитальности объекта дана в документах собственника': 0,
                'выбрать фактическую группу капитальности': 1,
            }
            viborka = st.selectbox(
                'фактичсекая капитальность объекта',
                vibor_fact_group.keys(), key=str(klfg)+'5432542555'
            )
            vibor = vibor_fact_group.get(viborka)
            if vibor == 0:
                vbrn_kap= int(st.number_input('введите фактическую капитальность объекта',min_value =1, step = 1, key=str(klfg)+'5565363265'))
            else:
                vbrn_kap= int(st.number_input('введите фактическую капитальность объекта, определив её по предложенной таблице',min_value =1, step = 1, key=str(klfg)+'5563489'))
                
                if straka_D <=6:
                    st.image('fact_kap3.png')
                else:
                    st.image('fact_kap1.png')
            if A == vbrn_kap:
                k3 = 1
            elif A == 0:
                k3 = 1
            else:
                if straka_D <=6:
                    book_V4 = op.open('В4.xlsx', read_only=True)
                    start_V4 = book_V4.active
                    vc_V4 = dict()
                    for ivv, rowvv in enumerate(range(1,start_V4.max_row +1)):
                        sub_rf_V4 = start_V4[rowvv][0].value
                        vc_V4[sub_rf_V4] = ivv

                    sub_russia_V4 = st.selectbox(
                        'Объект строительства',
                        vc_V4, key=str(klfg)+'588836345'
                    )
                                            
                    for dvv,qvv in enumerate(range(1,start_V4.max_row +1)):
                        if sub_russia_V4 == start_V4[qvv][0].value:
                            line_V4 = qvv

                    k3 = start_V4[line_V4][vbrn_kap].value
                else:
                    book_V2 = op.open('В2.xlsx', read_only=True)
                    start_V2 = book_V2.active
                    vc_V2 = dict()
                    for iv, rowv in enumerate(range(1,start_V2.max_row +1)):
                        sub_rf_V2 = start_V2[rowv][0].value
                        vc_V2[sub_rf_V2] = iv

                    sub_russia_V2 = st.selectbox(
                        'Объект строительства',
                        vc_V2, key=str(klfg)+'5888356346'
                    )
                                            
                    for dv,qv in enumerate(range(1,start_V2.max_row +1)):
                        if sub_russia_V2 == start_V2[qv][0].value:
                            line_V2 = qv

                    k3 = start_V2[line_V2][vbrn_kap].value
    #закончили к3 
    # начали коэф повреждения 
            if klfb in range(1,5):
                kf_an_isn = 0.35
            elif klfb == 5:
                kf_an_isn = 0.24
            elif klfb in range(6,14):
                kf_an_isn = 0.37
            elif klfb == 14:
                kf_an_isn = 0.66
            elif vibor_A == 25:
                kf_an_isn = 0.66
            elif line_A in range(18,24):
                kf_an_isn = 0.67
            else:
                kf_an_isn = 0.64
            Y1 = (kf_an_isn*C1*k2*k3*k4) + (C1*(1 - (K1/100))*k2*k3*k4)
        return Y1

    for klfg in range(0,Number):
        Y32_1 = Y_34_price(klfg)
        Y34.append(Y32_1)

    dfghj = sum(Y34)
    return dfghj
def y13():
    dict_y13 = []
    st.title('ущерб, нанесенный имуществу на объектах строительства, руб.')
    number_damage_im = int(st.number_input("введите кол-во Объектов, которым нанесён ущерб вследствие пожара", step = 1))
    
    def y13_11(w):
        fg = w+1
        st.write(f'Ущерб, нанесенный имуществу на объекте  {fg} типа')
        k_un = {"Здание производственного назначения":1,
        "Складское здание, сооружение (в т.ч. сельскохозяйственное здание, сооружение для хранения)":2,
        "Одноквартирный жилой дом":3,
        "Многоквартирный жилой дом":4,
        "Другое здание, сооружение, строение жилого сектора (кроме забора)":5,
        "Здание, сооружение сельскохозяйственного назначения (кроме зданий, сооружений для хранения)":6,
        "Здание, сооружение общественного назначения":7,
        "Другое здание, сооружение, строение (кроме сооружений, установок промышленного назначения, неэксплуатируемых, строящихся зданий, сооружений)":8,
        "иной вид объекта":0,
        }
        un_im = st.selectbox(
            'Вид объекта строительства',
            k_un.keys(), key=str(w)+'823427'
            )
        price_un_im = k_un.get(un_im)
        s1 = st.number_input('введите уничтоженную пожаром площадь объекта строительства, кв.м. определяемую в соответствие с порядком заполнения и представления КУП', key=str(w)+'83453457')
        st.write('В случае, если отсутствует уничтоженная пожаром площадь i-го объекта строительства, то приравнивается к площади пожара.')
        s3 = st.number_input('введите поврежденную пожаром площадь i-го объекта строительства, кв.м. определяемую в соответствии с Порядком заполнения и представления КУП', key=str(w)+'836367')   
        st.write('В случае если отсутствует поврежденная пожаром площадь /-го объекта строительства, то приравнивается к площади пожара.')
        if price_un_im == 1:
            y16 = 12686
            y25 = 9698
            y14 = y16*s1
            y15 = y25*s3
        elif price_un_im == 2:
            y16 = 24586
            y25 = 9840
            y14 = y16*s1
            y15 = y25*s3
        elif price_un_im == 3:
            y16 = 14386
            y25 = 3935
            y14 = y16*s1
            y15 = y25*s3
        elif price_un_im == 4:
            y16 = 12946
            y25 = 9271
            y14 = y16*s1
            y15 = y25*s3
        elif price_un_im == 5:
            y16 = 6078
            y25 = 5301
            y14 = y16*s1
            y15 = y25*s3
        elif price_un_im == 6:
            y16 = 32200
            y25 = 8023
            y14 = y16*s1
            y15 = y25*s3
        elif price_un_im == 7:
            y16 = 29439
            y25 = 27739
            y14 = y16*s1
            y15 = y25*s3
        elif price_un_im == 8:
            y16 = 17010
            y25 = 12763
            y14 = y16*s1
            y15 = y25*s3
        else:
            y16 = st.number_input('введите общую стоимость имущества, находившегося на момент пожара на i-м объекте строительства, руб.',min_value =0.01, key=str(w)+'8457')
            s2 = st.number_input('введите общую площадь объекта строительства, кв.м. определяемую в соответствии с Порядком заполнения и представления КУП',min_value =0.01, key=str(w)+'87')
            y14 = (y16/s2)*s1
            y15 = (0.75*(y16/s2))*s1
        y1311 = y14 + y15
        return y1311
    for w in range(0,number_damage_im):
        y133 = y13_11(w)
        dict_y13.append(y133)
    summa_y13 = sum(dict_y13)

    return summa_y13
def y26():
    y29 =[]
    st.title('Ущерб нанесённый пожаром транспортным средствам, руб.')
    number_un_ts = int(st.number_input("введите кол-во Т/С, которму нанесён ущерб ", step = 1))
    def y231(kraft):
        kraft_1 = kraft +1
        st.write(f'Рассчёт ущерба {kraft_1} транспортного средства')
        book = op.open('Н.xlsx', read_only=True) # начался просчёт коэфа 8
        staff = book.active
        vc = dict()
        for i, row in enumerate(range(1,staff.max_row +1)):
            sub_rf = staff[row][0].value
            vc[sub_rf] = i
        sub_russia = st.selectbox(
            'Субъекты рф',
         vc, key=str(kraft)+'252544'
        )
        for d,q in enumerate(range(1,staff.max_row +1)):
            if sub_russia == staff[q][0].value:
                liine = q
        k8 = staff[liine][1].value #вывели коэф 8 

    # начали си4
        st.write('Если в списке отсутствует вид вашего т/с, то выберите пункт "Иное Т/С"')
        file_XL = op.open('Л1.xlsx', read_only=True) # начали си4 
        status_file = file_XL.active
        dict_pril_L1 = dict()
        for z, w in enumerate(range(2,status_file.max_row +1)):
            vid_avto = status_file[w][0].value
            dict_pril_L1[vid_avto] = z
        vids_avto = st.selectbox(
            'вид т/с ',
            dict_pril_L1, key=str(kraft)+'22244'
        )
        dict_pril_L2 = dict()
        for hj,fg in enumerate(range(1,status_file.max_column)):
            otech_imp = status_file[1][fg].value
            dict_pril_L2[otech_imp] = hj
        vids_avto_2 = st.selectbox(
            'разновидности моделей т/с ',
            dict_pril_L2, key=str(kraft)+'76544'
        )
        for x,m in enumerate(range(2,status_file.max_row +1)):
            if vids_avto == status_file[m][0].value:
                straka = m
        for h,y in enumerate(range(1,status_file.max_column)):
             if vids_avto_2 == status_file[1][y].value:
                hex = y 
        if straka == 79:
            c4 = st.number_input('введите стоимость i-го транспортного средства в кв. м.',key=str(kraft)+'4544544')
        else:
            c4 = status_file[straka][hex].value
    #закончили си4 
        if straka in range(65, 72) or straka in range(2, 26):  # начали к7
            M1 = {
                "0-5 лет":0.8,
                "6-10 лет": 0.6,
                "11-15 лет": 0.5,
                "16-20 лет":0.35,
                "более 21 года":0.3,
            }
            M1_kf = st.selectbox(
                'срок эксплуатациитранспортного средства на момент пожара',
                M1.keys(), key=str(kraft)+'43664'
            )
            k7 = M1.get(M1_kf)
        elif straka in range(58, 65) or straka in range(72, 79):
            M2 = {
                "0-5 лет":0.8,
                "6-10 лет": 0.65,
                "11-15 лет": 0.55,
                "16-20 лет":0.4,
                "более 21 года":0.35,
            }
            M2_kf = st.selectbox(
                'срок эксплуатациитранспортного средства на момент пожара',
                M2.keys(), key=str(kraft)+'44567764'
            )
            k7 = M2.get(M2_kf)
        else:
            k7 = 1     #закончили к7
    #начали к9 и к10    
        if straka in range(58, 64) and hex == 1:  
            first_step_kf = 1
        elif straka in range(58, 64) and hex == 2:
            first_step_kf = 2
        elif straka in range(64,67):
            first_step_kf = 3
        elif straka in range(19,26):
            first_step_kf = 4
        elif straka in range(2,19):
            first_step_kf = 5
        elif straka in range(26,30):
            first_step_kf = 6
        elif straka in range(30,58) or straka in range(67,72):
            first_step_kf = 7
        else:
            first_step_kf = 8

        s14 = st.number_input('введите общую площадь i-го транспортного средства в кв. м.',min_value = 1, key=str(kraft)+'443544544')
        rule_for_s14 = {
            'Транспортное средство повреждено': 1,
            'Транспортное средство уничтоженно': 2,
        }
        list_for_rules_s14 = st.selectbox(
            'Выбериите Поврежденна или уничтожена сгоревшая площадь т/с',
            rule_for_s14.keys(), key=str(kraft)+'44978557764'
        )
        kf_rules = rule_for_s14.get(list_for_rules_s14) 
    
    
        if kf_rules == 1:
            k9 = 0
            s4 = 0
            s5 = st.number_input('введите повреждённую в результате горения площадь i-го т/с. в кв. м.',min_value =1, key=str(kraft)+'46786794')
            if first_step_kf == 1:
                if s14 <= 7:
                    if s5 in range(0, 2):
                        k10 = 0.53
                    elif s5 in range(2,3):
                        k10 = 0.79
                    elif s5 in range(3, 7):
                        k10 = 1
                else:
                    if s5 in range(0, 2):
                        k10 = 0.29
                    elif s5 in range(2,3):
                        k10 = 0.44
                    elif s5 in range(3, 7):
                        k10 = 0.55
                    else:
                        k10 = 1
            elif first_step_kf == 2:
                if s5 in range(0, 2):
                    if s14 in range(0, 2):
                        k10 = 0.66
                    elif s14 in range(2, 5):
                        k10 = 0.47
                    elif s14 in range(5,8):
                        k10 = 0.43
                    else:
                        k10 = 0.33
                elif s5 in range(2,3):
                    if s14 in range(2, 5):
                        k10 = 0.73
                    elif s14 in range(5,8):
                        k10 = 0.66
                    else:
                        k10 = 0.51
                else:
                    if s14 in range(2, 5):
                        k10 = 1
                    elif s14 in range(5,8):
                        k10 = 0.92
                    else:
                        k10 = 0.7
            elif first_step_kf == 3:
                if s14 <= 7:
                    if s5 in range(0, 2):
                        k10 = 0.28
                    elif s5 in range(2,3):
                        k10 = 0.42
                    elif s5 in range(3,5):
                        k10 = 0.53   
                    elif s5 in range(5,6):
                        k10 = 0.71
                    elif s5 in range(6,12):
                        k10 = 0.92
                elif s14 in range(8, 10):
                    if s5 in range(0, 2):
                        k10 = 0.23
                    elif s5 in range(2,3):
                        k10 = 0.35
                    elif s5 in range(3,5):
                        k10 = 0.43   
                    elif s5 in range(5,6):
                        k10 = 0.58
                    elif s5 in range(6,12):
                        k10 = 0.76
                elif s14 in range(10, 13):
                    if s5 in range(0, 2):
                        k10 = 0.2
                    elif s5 in range(2,3):
                        k10 = 0.31
                    elif s5 in range(3,5):
                        k10 = 0.38   
                    elif s5 in range(5,6):
                        k10 = 0.51
                    elif s5 in range(6,12):
                        k10 = 0.67
                    else: 
                        k10 = 0.79
                elif s14 in range(13, 20):
                    if s5 in range(0, 2):
                        k10 = 0.12
                    elif s5 in range(2,3):
                        k10 = 0.19
                    elif s5 in range(3,5):
                        k10 = 0.23   
                    elif s5 in range(5,6):
                        k10 = 0.31
                    elif s5 in range(6,12):
                        k10 = 0.4
                    else: 
                        k10 = 0.48
                elif s14 == 20:
                    if s5 in range(0, 2):
                        k10 = 0.08
                    elif s5 in range(2,3):
                        k10 = 0.12
                    elif s5 in range(3,5):
                        k10 = 0.15   
                    elif s5 in range(5,6):
                        k10 = 0.21
                    elif s5 in range(6,12):
                        k10 = 0.27
                    else: 
                        k10 = 0.32
                else:
                    if s5 in range(0, 2):
                        k10 = 0.07
                    elif s5 in range(2,3):
                        k10 = 0.1
                    elif s5 in range(3,5):
                        k10 = 0.13   
                    elif s5 in range(5,6):
                        k10 = 0.17
                    elif s5 in range(6,12):
                        k10 = 0.22
                    elif s5 in range(12,21): 
                        k10 = 0.26
                    else:
                        k10 = 0.65
            elif first_step_kf == 4:
                if s14 in range(0,3):
                    if s5 in range(0, 2):
                        k10 = 0.52
                    else:
                        k10 = 0.7
                elif s14 in range(3,4):
                    if s5 in range(0, 2):
                        k10 = 0.32
                    elif s5 in range(2,3):
                        k10 = 0.43
                    else:
                        k10 = 0.79
                elif s14 in range(4,5):
                    if s5 in range(0, 2):
                        k10 = 0.26
                    elif s5 in range(2,3):
                        k10 = 0.35
                    elif s5 in range(3,4):
                        k10 = 0.63
                    else:
                        k10 = 0.83
                elif s14 in range(5,19):
                    if s5 in range(0, 2):
                        k10 = 0.13
                    elif s5 in range(2,3):
                        k10 = 0.18
                    elif s5 in range(3,4):
                        k10 = 0.32
                    elif s5 in range(4,5):
                        k10 = 0.42
                    elif s5 in range(5,6):
                        k10 = 0.5
                    else: 
                        k10 = 0.78 
                else:
                    if s5 in range(0, 2):
                        k10 = 0.08
                    elif s5 in range(2,3):
                        k10 = 0.11
                    elif s5 in range(3,4):
                        k10 = 0.2
                    elif s5 in range(4,5):
                        k10 = 0.26
                    elif s5 in range(5,6):
                        k10 = 0.3
                    elif s5 in range(6,20): 
                        k10 = 0.48
                    else:
                        k10 = 0.81
            elif first_step_kf == 5:
                if s14 in range(0,6):
                    if s5 in range(0,6):
                        k10 = 0.55
                elif s14 in range(6,7):
                    if s5 in range(0,6):
                        k10 = 0.29
                    else:
                        k10 = 0.41
                elif s14 in range(7,12):
                    if s5 in range(0,6):
                        k10 = 0.2
                    elif s5 in range(6,10):
                        k10 = 0.27
                    else:
                        k10 = 0.52
                else:
                    if s5 in range(0,6):
                        k10 = 0.17
                    elif s5 in range(6,10):
                        k10 = 0.23
                    elif s5 in range(10,12):
                        k10 = 0.44
                    else:
                        k10 = 0.65
            elif first_step_kf == 6:
                if s14 in range(0,6):
                    if s5 in range(0,4):
                        k10 = 0.39
                    else:
                        k10 = 1
                elif s14 in range(6,20):
                    if s5 in range(0,4):
                        k10 = 0.24
                    elif s5 in range(4,6):
                        k10 = 0.61
                    else:
                        k10 = 0.85
                else:
                    if s5 in range(0,4):
                        k10 = 0.07
                    elif s5 in range(4,6):
                        k10 = 0.17
                    elif s5 in range(6,21):
                        k10 = 0.24
                    else:
                        k10 = 0.77
            elif first_step_kf == 7:
                if s14 in range(0,6):
                    if s5 in range(0,5):
                        k10 = 0.47
                    else:
                        k10 = 0.65
                elif s14 in range(7,8):
                    if s5 in range(0,5):
                        k10 = 0.41
                    else:
                        k10 = 0.56
                else:
                    if s5 in range(0,5):
                        k10 = 0.36
                    elif s5 in range(6,15):
                        k10 = 0.49
                    elif s5 in range(15,20):
                        k10 = 0.92
                    else:
                        k10 = 1
            else:
                if s14 in range(0,2):
                    if s5 in range(0,2):
                        k10 = 0.97
                elif s14 in range(2,3):
                    if s5 in range(0,2):
                        k10 =0.64
                    else:
                        k10 = 0.73
                else:
                    if s5 in range(0,2):
                        k10 =0.48
                    else:
                        k10 = 0.55
        elif kf_rules == 2:
            k10 = 0
            s5 = 0
            s4 = st.number_input('введите уничтоженную в результате горения площадь i-го т/с. в кв. м.',min_value =1, key=str(kraft)+'42454554')
            if first_step_kf == 1:
                if s14 in range(0,7):
                    k9 = 1
                else: 
                    if s4 in range(0,7):
                        k9 = 0.55
                    else:
                        k9 = 1
            elif first_step_kf == 2:
                if s14 in range(0,2):
                    k9 = 1
                elif s14 in range(2,5):
                    if s4 in range(0,2):
                        k9 = 0.71
                    else:
                        k9 = 1
                elif s14 in range(5,8):
                    if s4 in range(0,2):
                        k9 = 0.65
                    elif s4 in range(2,5):
                        k9 = 0.92
                    else:
                        k9 = 1
                else:
                    if s4 in range(0,2):
                        k9 = 0.5
                    elif s4 in range(2,5):
                        k9 = 0.7
                    elif s4 in range(5,8):
                        k9 = 0.77
                    else:
                        k9 = 1
            elif first_step_kf == 3:
                if s14 in range(0,8):
                    k9 = 1
                elif s14 in range(8,10):
                    if s4 in range(0,8):
                        k9 = 0.82
                    else:
                        k9 = 1
                elif s14 in range(10,13):
                    if s4 in range(0,8):
                        k9 = 0.72
                    elif s4 in range(8,10):
                        k9 = 0.88
                    else:
                        k9 = 1
                elif s14 in range(13,20):
                    if s4 in range(0,8):
                        k9 = 0.44
                    elif s4 in range(8,10):
                        k9 = 0.53
                    elif s4 in range(10,13):
                        k9 = 0.6
                    else:
                        k9 = 1
                elif s14 in range(20,21):
                    if s4 in range(0,8):
                        k9 = 0.29
                    elif s4 in range(8,10):
                        k9 = 0.35
                    elif s4 in range(10,13):
                        k9 = 0.4
                    elif s4 in range(13,20):
                        k9 = 0.66
                    else:
                        k9 = 1
                else:
                    if s4 in range(0,8):
                        k9 = 0.24
                    elif s4 in range(8,10):
                        k9 = 0.29
                    elif s4 in range(10,13):
                        k9 = 0.33
                    elif s4 in range(13,20):
                        k9 = 0.55
                    elif s4 in range(20,21):
                        k9 = 0.82
                    else:
                        k9 = 1
            elif first_step_kf == 4:
                if s14 in range(0,3):
                    k9 = 1 
                elif s14 in range(3,4):
                    if s4 in range(0,3):
                        k9 = 0.62
                    else:
                        k9 =1 
                elif s14 in range(4,5):
                    if s4 in range(0,3):
                        k9 = 0.5
                    elif s4 in range(3,4):
                        k9 = 0.8
                    else: 
                        k9 =1
                elif s14 in range(5,19):
                    if s4 in range(0,3):
                        k9 = 0.25
                    elif s4 in range(3,4):
                        k9 = 0.41
                    elif s4 in range(4,5): 
                        k9 = 0.51
                    else:
                        k9 = 1
                else:
                    if s4 in range(0,3):
                        k9 = 0.15
                    elif s4 in range(3,4):
                        k9 = 0.25
                    elif s4 in range(4,5): 
                        k9 = 0.31
                    elif s4 in range(5,19):
                        k9 = 0.61
                    else:
                        k9 =1
            elif first_step_kf == 5:
                if s14 in range(0,6):
                    k9 = 1
                elif s14 in range(6,7):
                    if s4 in range(0,6):
                        k9 = 0.53
                    else:
                        k9 =1
                elif s14 in range(7,12):
                    if s4 in range(0,6):
                        k9 = 0.36
                    elif s4 in range(6,7):
                        k9 = 0.67
                    else:
                        k9 = 1
                else:
                    if s4 in range(0,6):
                        k9 = 0.3
                    elif s4 in range(6,7):
                        k9 = 0.57
                    elif s4 in range(7,12):
                        k9 = 0.85
                    else:
                        k9 = 1
            elif first_step_kf == 6:
                if s14 in range(0,6):
                    k9 = 1
                elif s14 in range(6,20):
                    if s4 in range(0,6):
                        k9 = 0.61
                    else:
                        k9 = 1
                else:
                    if s4 in range(0,6):
                        k9 = 0.17
                    elif s4 in range(6,20):
                        k9 = 0.28
                    else:
                        k9 = 1
            elif first_step_kf == 7:
                if s14 in range(0,7):
                    k9 = 1
                elif s14 in range(7,9):
                    if s4 in range(0,7):
                        k9 = 0.86
                    else:
                        k9 = 1
                else:
                    if s4 in range(0,7):
                        k9 = 0.76
                    elif s4 in range(7,9):
                        k9 = 0.88
                    else:
                        k9 = 1
            else:
                if s14 in range(0,2):
                    k9 = 1
                elif s14 in range(2,3):
                    if s4 in range(0,2):
                        k9 = 0.66
                    else:
                        k9 = 1
                else:
                    if s4 in range(0,2):
                        k9 = 0.49
                    elif s4 in range(2,3):
                        k9 = 0.75
                    else:
                        k9 = 1
    #закончили к9 и к10
        if kf_rules == 1:
            y28 = c4*(1 + (k7 - 1)*k8)*(k9 + k10*(1 - k9))
            damage = y28
        elif straka == 79:
            c5 = c4/s14
            y28 = c5*(s4+0.94*s5)
            damage =y28
        else:
            y27 = c4*(1 + (k7 - 1)*k8)
            damage = y27

        return damage
    for kraft in range(0, number_un_ts):
        y2_31 = y231(kraft)
        y29.append(y2_31)
    zcvzb = sum(y29)
    return zcvzb
def y9_summa():
    y9 = []
    st.title('ущерб, нанесенный имуществу на транспортных средствах, руб.')
    Number_ucherb_im_ts =int(st.number_input("введите кол-во имущества, которму нанесён ущерб на трансопртных средствах", step = 1))

    def input_data(i):
        g = i+1
        st.write(f'ущерб, нанесенный {g} имуществу на транспортных средствах')
        s7 = st.number_input("введите общую площадь имущества, находившегося на момент пожара на г-м транспортном средстве, кв.м.",min_value =0.01, key=str(i)+'12')
        c7 =  st.number_input("введите общую стоимость имущества, находившегося на момент пожара на г-м транспортном средстве, руб.",min_value =0.01, key=str(i)+'13')
        s6 =  st.number_input("введите площадь уничтоженного имущества на г-м транспортном средстве, кв.м.",min_value =0.01, key=str(i)+'14')
        s8 =  st.number_input("введите площадь поврежденного имущества на г-м транспортном средстве, кв.м.",min_value =0.01, key=str(i)+'15')

        def y17():
            c6 = c7/s7

            y18 = c6*s6
            y19 = 0.94*c6*s8
            y1711 = y18 + y19 
            return y1711
        damage = y17()  
        return damage
    for i in range(0,Number_ucherb_im_ts):    
        y9_1 = input_data(i)
        y9.append(y9_1)

    vbnm = sum(y9)

    return vbnm
def y10():
    price_list_cost = []
    st.title('ущерб в результате уничтожения сельскохозяйственных посевов, лесных насаждений, руб.')
    number_get_costs = int(st.number_input("Введите кол-во видов уничтоженных сельскохозяйственных посевов, лесных насаждений", step = 1))
    def get_cost(r):
        fd = r+1 
        st.write(f'ущерб, нанесенный уничтоженному сельскохозяйственному посеву, лесному насаждению {fd} вида')
        costs = {
            "Соя": 4.63,
            "Рапс": 5.87,
            "Рожь":  2.56,
            "Овес": 1.63,
            "Просо": 1.40,
            "Ячмень":   3.00,
            "Гречиха": 3.25,
            "Пшеница": 3.75,
            "Кукуруза": 5.99,
            "Зернобобовые": 2.73,
            "Подсолнечник": 4.23,
            "Однолетние травы": 0.95,
            "Многолетние травы": 1.04,
            "Лен - долгунец (волокно)": 3.18,
            "Прочие зерновые и зернобобовые": 3.69,
            "плодово-ягодные, овощно-бахчевые": 126.63,
            "Иная сельхозяйственная культура": 0,
        }
        plant = st.selectbox(
            'Пострадала сельхозяйственная культура',
            costs.keys(), key=str(r)+'71347'
        )
        st.write('Введите:')
        
        destroyed_area = st.number_input('Общую уничтоженную пожаром площадь г-х сельскохозяйственных посевов (лесного насаждения), кв.м.',min_value =0.01, key=str(r)+'7356367')
        price = costs.get(plant) 
        if price == 0:
            total_area = st.number_input('Общую площадь i-х сельскохозяйственных посевов (лесного насаждения), кв.м.',min_value =0.01, key=str(r)+'76347')
            total_cost = st.number_input('Общую стоимость i-х сельскохозяйственных посевов (лесного насаждения), руб.',min_value =0.01, key=str(r)+'363677')
            res = (total_cost/total_area)*destroyed_area
        else:	
            res = price*destroyed_area
        return res
    for r in range(0, number_get_costs):
        get_cost_v = get_cost(r)
        price_list_cost.append(get_cost_v)
    price_one_get_cost = sum(price_list_cost)

    return price_one_get_cost
def y21():
    y21_tabl =[]
    st.title('ущерб, нанесенный имуществу на открытой территории, руб.')
    Number_price_open_ter =int(st.number_input("введите кол-во имущества, которму нанесён ущерб на открытой территории", step = 1))
    def y21_1(e):
        f = e +1 
        st.write(f'Рассчёт ущерба нанесенный {f} имуществу на открытой территории')
        s12 = st.number_input("введите общую площадь имущества i-го вида на открытой территории, кв.м",min_value =0.01, key=str(e)+'8258')
        c11 = st.number_input("введите общую стоимость имущества i-го вида на открытой территории, руб.",min_value =0.01, key=str(e)+'25588')
        s11 = st.number_input("введите площадь уничтоженного имущества i-го вида на открытой территории, кв.м.",min_value =0.01, key=str(e)+'823558')
        s13 = st.number_input("введите площадь поврежденного имущества i-го вида на открытой территории, кв.м.",min_value =0.01, key=str(e)+'823523568')
        c10 = c11/s12
        y22 = (c10*s11)
        y23 = (0.75*c10*s13)

        y2111 = y22 + y23
        return y2111
    for e in range(0, Number_price_open_ter):    
        y21_11 =  y21_1(e)
        y21_tabl.append(y21_11)

    vbgh = sum(y21_tabl)

    return vbgh
def y12():
    list_death = []
    st.title("Ущерб, нанесенный животным в руб.")
    Number_death_animal = int(st.number_input("Введите кол-во видов уничтоженных животных", step = 1))
    def get_animal(q):
        mm = q+1 
        st.write(f'Ущерб, нанесенный животным {mm} типа')
        price_animals_list = {
            "Корова":66545,
            "Бык":114941,
            "Телёнок, бычок(годовалые)":30248,
            "Овца":7331,
            "Баран":9775,
            "Овца, Баран(годовалые)":4887,
            "Коза":4887,
            "Козел":6109,
            "Свинья":19990,
            "Кабан":29986,
            "Поросёнок (годовалый)":17991,
            "Осел":28291,
            "Мул":37721,
            "Пони":18149,
            "Верховая (легкоупряжная) лошадь":54446,
            "Тяжеловозная лошадь":102842,
            "Северный олень (самка)":15729,
            "Северный олень (самец)":20568,
            "Курица (петух)":273,
            "Гусь":642,
            "Утка":401,
            "Индюк":1283,
            "Индюшка":722,
            "Кролик":3600,
            "Пчелосемья":4000,
        } 
        animal = st.selectbox(
            'Выберите животное',
            price_animals_list.keys(),key=str(q)+'133300'
        )
        number_animals =st.number_input('количество животных данного вида', step = 1,key=str(q)+'10022')

        price_animal = price_animals_list.get(animal)
        res_animal = price_animal*number_animals
        return res_animal 
    for q in range(0, Number_death_animal):
        y12_1 = get_animal(q)
        list_death.append(y12_1)

    summ_price = sum(list_death)
    return summ_price

list_price = []
st.title('Расчет материального ущерба,нанесеного пожаром объектам строительства и имуществу')
p1 = st.checkbox('Ущерб нанесённый пожаром объектам строительства')
if p1:
    k1 =Y1()
    list_price.append(k1)
p2 = st.checkbox('ущерб, нанесенный имуществу на объектах строительства')
if p2:
    k2 =y13()
    list_price.append(k2)
p3 = st.checkbox('Ущерб нанесённый пожаром транспортным средствам')
if p3:
    k3 =y26()
    list_price.append(k3)
p4 = st.checkbox('ущерб, нанесенный имуществу на транспортных средствах')
if p4:
    k4 = y9_summa()
    list_price.append(k4)
p5 = st.checkbox('ущерб в результате уничтожения сельскохозяйственных посевов, лесных насаждений')
if p5:
    k5 =y10()
    list_price.append(k5)
p6 = st.checkbox('ущерб, нанесенный имуществу на открытой территории')
if p6:
    k6=y21()
    list_price.append(k6)
p7 = st.checkbox('Ущерб, нанесенный животным')
if p7:
    k7=y12()
    list_price.append(k7)
damage = sum(list_price)  

if st.button('Вычислить..'):
    st.subheader('Результат:')
    st.write(damage)
st.image('SPSA2.png')
